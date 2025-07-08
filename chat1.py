import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment, Border, Side


def file_xl(file_name, data1, data2, data3):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'x1'

    # تلوين الخلفية كلها أبيض
    for i in range(1, 51):
        for j in range(1, 12):
            fill_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ws.cell(i, j).fill = fill_color

    # إدراج اللوجو
    img = Image('Mo.png')
    img.width = 793
    img.height = 190
    ws.add_image(img, 'A1')

    # تلوين خلفية بيانات seller / bill to
    for i in range(11, 19):
        for j in range(1, 12):
            fill_color = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws.cell(i, j).fill = fill_color

    # تعبئة بيانات البائع data1
    data1_keys = list(data1.keys())
    for index, i in enumerate(range(13, 17)):
        ws.cell(i, 2).value = data1_keys[index]
        ws.cell(i, 3).value = ": " + data1[data1_keys[index]]

    # تعبئة بيانات المشتري data2
    data2_keys = list(data2.keys())
    for index, i in enumerate(range(13, 17)):
        ws.cell(i, 6).value = data2_keys[index]
        ws.cell(i, 7).value = ": " + data2[data2_keys[index]]

    # ترويسة الجدول
    header_row = 20
    columns = {
        'B': "No.",
        'D': "Description",
        'F': "Quantity",
        'H': "Item Price",
        'J': "Total"
    }

    for col, title in columns.items():
        cell = f"{col}{header_row}"
        ws[cell].value = title
        ws[cell].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # تعبئة صفوف الجدول من data3
    start_row = header_row + 1
    for idx, key in enumerate(data3.keys(), start=start_row):
        item = data3[key]
        ws.cell(idx, 2).value = str(idx - header_row)  # No.
        ws.cell(idx, 4).value = item['Description']
        ws.cell(idx, 6).value = item['Quantity']
        ws.cell(idx, 8).value = item['Item Price']
        ws.cell(idx, 10).value = item['Total']

        for j in range(2, 11, 2):
            ws.cell(idx, j).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # حساب الإجماليات
    end_row = start_row + len(data3)
    subtotal = sum(int(data3[key]['Total']) for key in data3)
    tax = int(subtotal * 0.10)
    grand_total = subtotal + tax

    ws.cell(end_row, 9).value = "Subtotal"
    ws.cell(end_row, 10).value = subtotal

    ws.cell(end_row + 1, 9).value = "Tax (10%)"
    ws.cell(end_row + 1, 10).value = tax

    ws.cell(end_row + 2, 9).value = "Grand Total"
    ws.cell(end_row + 2, 10).value = grand_total

    # ملاحظات
    note_row = end_row + 5
    ws.cell(note_row, 2).value = "Notes:"
    ws.cell(note_row + 1, 2).value = "1. Payment is due within 30 days from the date of the invoice."
    ws.cell(note_row + 2, 2).value = "2. Please make payment to the following bank account"

    ws.cell(note_row + 4, 2).value = "Bank Name"
    ws.cell(note_row + 4, 3).value = ": Bank XYZ"
    ws.cell(note_row + 5, 2).value = "Account Number"
    ws.cell(note_row + 5, 3).value = ": 123-456-789"
    ws.cell(note_row + 6, 2).value = "Account Holder"
    ws.cell(note_row + 6, 3).value = ": X Construction"

    # حفظ الملف
    wb.save(file_name + '.xlsx')


# بيانات الإدخال
data1 = {
    "Seller": "X Construction",
    "Address": "873 Liberty Street, Las Vegas",
    "Mail": "xconstruction@mail.com",
    "Phone": "+1 312-692-0767",
}

data2 = {
    "Bill To": "ABC Company",
    "Address": "123 Main Street Cityville",
    "Mail": "abc@mail.com",
    "Phone": "+1 312-483-8673",
}

data3 = {
    "Foundation Work": {
        "Description": "Foundation Work",
        "Quantity": "10 Days",
        "Item Price": "100",
        "Total": "1000"
    },
    "Steel Structure Installation": {
        "Description": "Steel Structure Installation",
        "Quantity": "5 Weeks",
        "Item Price": "2000",
        "Total": "10000"
    },
    "Concrete Material": {
        "Description": "Concrete Material",
        "Quantity": "200 Cubics",
        "Item Price": "50",
        "Total": "10000"
    },
    "Structural Steel Material": {
        "Description": "Structural Steel Material",
        "Quantity": "10 Tons",
        "Item Price": "500",
        "Total": "5000"
    },
}

# تنفيذ الدالة
file_xl('Mo', data1, data2, data3)
