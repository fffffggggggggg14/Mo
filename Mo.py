import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def file_xl(file_name, img, data1, data2, data3, data4, data5, rate):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'x1'

    data3_key = list(data3.keys())
    data3_values = list(data3[data3_key[0]].keys())
    # print(data3_key)
    # print(data3_values)

    key_data4 = list(data4.keys())
    print(key_data4)


    len1 = len(data3_key)
    len2 = len(data4[key_data4[0]])
    len3 = len(data5)
    
    h = len1 + len2 + len3 + 34

    print(h)

    for i in range(1, h):
        for j in range(1, 12):
            fill_color = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
            ws.cell(i, j).fill = fill_color

    img = Image(img)
    img.width = 793
    img.height = 190
    ws.add_image(img, 'A1')
    


    for i in range(11, 19):
        for j in range(1, 12):
            fill_color = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws.cell(i, j).fill = fill_color

    # for i in range(11, 19):
    #     for j in range(1, 12):
    #         ws.cell(i, j).border = xl.styles.Border(left=xl.styles.Side(style='thin'),
    #                                                 right=xl.styles.Side(style='thin'),
    #                                                 top=xl.styles.Side(style='thin'),
    #                                                 bottom=xl.styles.Side(style='thin'))
    
    
    ws.merge_cells('C13:E13')
    ws.merge_cells('C14:E14')
    ws.merge_cells('C15:E15')
    ws.merge_cells('C16:E16')

    data1_keys = list(data1.keys())

    # print(data1_keys)

    for index, i in enumerate(range(13, 17), start = 0):
        ws.cell(i, 2).value = data1_keys[index]
        ws.cell(i, 2).font = Font(bold = True)
    
    for index, i in enumerate(range(13, 17), start = 0):
        ws.cell(i, 3).value = ":" + data1[data1_keys[index]]


    
    ws.merge_cells('G13:I13')
    ws.merge_cells('G14:I14')
    ws.merge_cells('G15:I15')
    ws.merge_cells('G16:I16')

    data2_keys = list(data2.keys())

    # print(data2_keys)

    for index, i in enumerate(range(13, 17), start = 0):
        ws.cell(i, 6).value = data2_keys[index]
        ws.cell(i, 6).font = Font(bold = True)
    
    for index, i in enumerate(range(13, 17), start = 0):
        ws.cell(i, 7).value = ":" + data2[data2_keys[index]]



    for i in range(2, 11):
            fill_color = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            ws.cell(20, i).fill = fill_color


    for i in range(2, 11):
            ws.cell(20, i).border = xl.styles.Border(left=xl.styles.Side(style='thin'),
                                                    right=xl.styles.Side(style='thin'),
                                                    top=xl.styles.Side(style='thin'),
                                                    bottom=xl.styles.Side(style='thin'))        




    
    ws.merge_cells('C20:E20')
    ws.merge_cells('G20:H20')
    ws.merge_cells('I20:J20')

    ws['B20'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C20'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G20'].alignment = Alignment(horizontal='center', vertical='center')
    ws['I20'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(20, 2).value, ws.cell(20, 3).value, ws.cell(20, 6).value, ws.cell(20, 7).value, ws.cell(20, 9).value = ["No."] + data3_values

    ws.cell(20, 2).font = Font(bold = True, color = 'ffffff')
    ws.cell(20, 3).font = Font(bold = True, color = 'ffffff') 
    ws.cell(20, 6).font = Font(bold = True, color = 'ffffff')
    ws.cell(20, 7).font = Font(bold = True, color = 'ffffff')
    ws.cell(20, 9).font = Font(bold = True, color = 'ffffff')



    len1 = len(data3_key)
    for i in range(21, 21 + len1):
        ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{i}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'I{i}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells(f'C{i}:E{i}')
        ws.merge_cells(f'G{i}:H{i}')
        ws.merge_cells(f'I{i}:J{i}')
    
    

    for i in range(len1):   
        j = 0
        ws.cell(21 + i, 2).value = i + 1
        ws.cell(21 + i, 3).value = data3[data3_key[i]][data3_values[j]]
        j += 1
        ws.cell(21 + i, 6).value = data3[data3_key[i]][data3_values[j]]
        ws.cell(21 + i, 6).font = Font(size = 10)
        j += 1
        ws.cell(21 + i, 7).value = f"${int(data3[data3_key[i]][data3_values[j]]):,}"
        j += 1
        ws.cell(21 + i, 9).value = f"${int(data3[data3_key[i]][data3_values[j]]):,}"
        j = 0
    
    
    x = [2, 3, 4, 5, 6, 7, 8, 9, 10]
    for i in range(21, 21 + len1):
        for j in x:
            ws.cell(i, j).border = xl.styles.Border(
                left=xl.styles.Side(style='thin'),
                right=xl.styles.Side(style='thin'),
                top=xl.styles.Side(style='thin'),
                bottom=xl.styles.Side(style='thin')
            )

            subtotal = sum([int(v['Total']) for v in data3.values()])
            tax = int(subtotal * (rate / 100))
            grand_total = subtotal + tax

            values = [
                ("Subtotal", subtotal),
                (f"Tax ({rate}%)", tax),
                ("Grand Total", grand_total)
            ]

            start_row = 21 + len1
            start_col = 7 

            for idx, (label, val) in enumerate(values):
                row = start_row + idx

                ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)
                ws.merge_cells(start_row=row, start_column=start_col+2, end_row=row, end_column=start_col+3)

                ws.cell(row, start_col).value = label
                ws.cell(row, start_col+2).value = f"${val:,}"

                for col in [start_col, start_col+2]:
                    cell = ws.cell(row, col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )


    end = 21 + len1 + 4

    for i in range(2, 11):
        ws.cell(end, i).value = '- - - - - - - - -'
        ws.cell(end, i).font = Font(color = '1F4E78', bold = True)
        
    end += 2

    ws[f'B{end}'].value = key_data4[0] + ":"
    ws[f'B{end}'].font = Font(bold = True)

    end += 1

    len2 = len(data4[key_data4[0]])
    for index, i in enumerate(range(end, end + len2)):
        ws.cell(i, 2).value = f"{index + 1}. " + data4[key_data4[0]][index]
        ws.merge_cells(f"B{i}:H{i}")


    end += len2 + 1



    for key, value in dict(data5).items():
        ws.cell(end, 3).value = key
        ws.merge_cells(f'C{end}:D{end}')

        ws.cell(end, 5).value = ": " + value
        ws.merge_cells(f'E{end}:F{end}')
        end += 1


    end += 2



    ws[f'E{end}'].value = "Thank Youfor You Business"
    ws[f'E{end}'].font = Font(color = "1F4E78", bold = True)


    end += 2


    for i in range(end, end + 2):
        for j in range(1, 12):
            fill_color = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            ws.cell(i, j).fill = fill_color

    end += 1
    

    print(end)

    wb.save(file_name + '.xlsx')























data1 = {
    "Seller" : "XConstruction", 
    "Address" : "873 Liberty Street,Las Vegas",
    "Mail" : "xconstruction",
    "Phone" : "+1 312-692-0767",
}

data2 = {
    "Bill To" : "ABC Company", 
    "Address" : "123 Main Street Cityville",
    "Mail" : "abc @mail.com",
    "Phone" : "+1 312-483-8673",
}


data3 = {
    "AAAAAAS" : {
        "Description" : "Foundation Work",
        "Quantity" : "10 Day",
        "Item Price" : "100",
        "Total" : "1000"
    },
    "Foundation Work" : {
        "Description" : "Foundation Work",
        "Quantity" : "10 Day",
        "Item Price" : "100",
        "Total" : "1000"
    },
    "AAA" : {
        "Description" : "Foundation Work",
        "Quantity" : "10 Day",
        "Item Price" : "100",
        "Total" : "1000"
    },
    "Steel Sturcture Installation" : {
        "Description" : "Steel Sturcture Installation",
        "Quantity" : "5 Week",
        "Item Price" : "2000",
        "Total" : "10000"
    },
    "Contrete Material" : {
        "Description" : "Contrete Material",
        "Quantity" : "200 Cubics",
        "Item Price" : "50",
        "Total" : "10000"
    },
    "Structural Steel Material" : {
        "Description" : "Structural Steel Material",
        "Quantity" : "10 Tons",
        "Item Price" : "500",
        "Total" : "5000"
    },
    "XXX" : {
        "Description" : "Structural Steel Material",
        "Quantity" : "10 Tons",
        "Item Price" : "500",
        "Total" : "5000"
    },
}


data4 = {
    "Note" : [
        "payment is due within 30 days from the data of the invoice.",
        "please make payment to the following bank account."
        ]
}


data5 = {
    "Bank Name" : "BankXYZ",
    "Account Number" : "123-456-789",
    "Account Holder" : "XCountruction"
}







file_xl('Mo', 'Mo.png', data1, data2, data3, data4, data5, 9)



