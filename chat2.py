import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def file_xl(file_name, data1, data2, data3):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'x1'

    # تلوين الخلفية
    for i in range(1, 51):
        for j in range(1, 12):
            ws.cell(i, j).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # إدراج الصورة
    img = Image('Mo.png')
    img.width = 793
    img.height = 190
    ws.add_image(img, 'A1')

    # تلوين خلفية بيانات البائع والمشتري
    for i in range(11, 19):
        for j in range(1, 12):
            ws.cell(i, j).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # بيانات البائع
    for idx, key in enumerate(data1.keys()):
        row = 13 + idx
        ws.cell(row, 2).value = key
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 3).value = ": " + data1[key]

    # بيانات المشتري
    for idx, key in enumerate(data2.keys()):
        row = 13 + idx
        ws.cell(row, 6).value = key
        ws.cell(row, 6).font = Font(bold=True)
        ws.cell(row, 7).value = ": " + data2[key]

    # تنسيق الأعمدة
    widths = {'B': 5, 'C': 25, 'D': 25, 'E': 25, 'F': 2, 'G': 10, 'H': 10, 'I': 10, 'J': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # دمج رؤوس الأعمدة
    ws.merge_cells('C20:E20')
    ws.merge_cells('G20:H20')
    ws.merge_cells('I20:J20')

    # كتابة رؤوس الجدول
    ws['B20'] = "No."
    ws['C20'] = "Description"
    ws['G20'] = "Quantity"
    ws['I20'] = "Item Price"
    ws['K20'] = "Total"  # Total خارج الدمج

    # تنسيق رؤوس الجدول
    header_cells = ['B20', 'C20', 'G20', 'I20', 'K20']
    for cell in header_cells:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        ws[cell].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # تعبئة البيانات في صفوف الجدول
    start_row = 21
    for idx, (key, val) in enumerate(data3.items(), start=0):
        row = start_row + idx

        # كتابة البيانات
        ws.cell(row, 2).value = idx + 1  # No.
        ws.cell(row, 3).value = val['Description']
        ws.cell(row, 6).value = val['Quantity']
        ws.cell(row, 9).value = val['Item Price']
        ws.cell(row, 11).value = val['Total']

        # دمج الخانات
        ws.merge_cells(f'C{row}:E{row}')
        ws.merge_cells(f'G{row}:H{row}')
        ws.merge_cells(f'I{row}:J{row}')

        # محاذاة
        for col in ['B', 'C', 'G', 'I', 'K']:
            cell = f'{col}{row}'
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')

        # حدود
        for col in [2, 3, 6, 8, 9, 11]:
            ws.cell(row, col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # حفظ الملف
    wb.save(file_name + '.xlsx')


# البيانات
data1 = {
    "Seller": "XConstruction",
    "Address": "873 Liberty Street, Las Vegas",
    "Mail": "xconstruction@mail.com",
    "Phone": "+1 312-692-0767",
}

data2 = {
    "Bill To": "ABC Company",
    "Address": "123 Main Street Cityville",
    "Mail": "abc@mail.com",
    "Phone": "+1 312-483-8673",
}

data3 = {
    "Foundation Work": {
        "Description": "Foundation Work",
        "Quantity": "10 Days",
        "Item Price": "100",
        "Total": "1000"
    },
    "Steel Structure Installation": {
        "Description": "Steel Structure Installation",
        "Quantity": "5 Weeks",
        "Item Price": "2000",
        "Total": "10000"
    },
    "Concrete Material": {
        "Description": "Concrete Material",
        "Quantity": "200 Cubics",
        "Item Price": "50",
        "Total": "10000"
    },
    "Structural Steel Material": {
        "Description": "Structural Steel Material",
        "Quantity": "10 Tons",
        "Item Price": "500",
        "Total": "5000"
    },
}

file_xl('Mo', data1, data2, data3)
